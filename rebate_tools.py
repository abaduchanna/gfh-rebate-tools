"""
Rebate Folder Tools
Runs three operations in sequence:
  1. Store Rename     — fix store name typos in filenames
  2. Add 2026 Suffix  — append " 2026" to Excel filenames that don't have it
  3. Delete 2025 Rows — remove rows where column A contains "2025" from .xls files
"""

import os
import sys
import time
import subprocess
import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox
from pathlib import Path
import threading

# ── Auto-install required packages BEFORE importing them ──────────────────
def _pip_install(pkg_name):
    subprocess.run(
        [sys.executable, "-m", "pip", "install", pkg_name,
         "--quiet", "--disable-pip-version-check"],
        capture_output=True,
    )

for _pkg, _pip in [
    ("openpyxl",  "openpyxl"),
    ("xlrd",      "xlrd==1.2.0"),
    ("xlwt",      "xlwt"),
    ("xlutils",   "xlutils"),
]:
    try:
        __import__(_pkg)
    except ImportError:
        _pip_install(_pip)

# ── Now import after ensuring packages are installed ──────────────────────
try:
    from openpyxl import load_workbook as _load_xlsx
except ImportError:
    _load_xlsx = None

try:
    import xlrd
    import xlwt
    from xlutils.copy import copy as xl_copy
except ImportError:
    xlrd = xlwt = xl_copy = None


# ═══════════════════════════════════════════════════════════════════════════
# STEP 1 — STORE RENAME
# ═══════════════════════════════════════════════════════════════════════════
STORE_RENAMES = [
    ("Eliff",       "E Iliff"),
    ("FM 1960",     "FM1960"),
    ("Mckellips",   "McKellips"),
    ("N 19th",      "N 19"),
    ("N 35th",      "N 35"),
    ("N 51st",      "N 51"),
    ("New Thomas",  "W Thomas"),
    ("Mt View RD",  "Mount View"),
    ("4363 W Fuqua St", "Fuqua"),
]

def step1_store_rename(folder: Path, log):
    log("\n── STEP 1: Store Rename ──────────────────────────────────")
    renamed = skipped = 0
    for f in sorted(folder.iterdir()):
        if not f.is_file():
            continue
        new_name = f.name
        for old, new in STORE_RENAMES:
            new_name = new_name.replace(old, new)
        if new_name != f.name:
            dest = f.parent / new_name
            try:
                f.rename(dest)
                log(f"  Renamed: {f.name}  →  {new_name}")
                renamed += 1
            except Exception as e:
                log(f"  ERROR renaming {f.name}: {e}")
        else:
            skipped += 1
    log(f"  Done — {renamed} renamed, {skipped} unchanged.")
    return renamed


# ═══════════════════════════════════════════════════════════════════════════
# STEP 2 — ADD " 2026" SUFFIX
# ═══════════════════════════════════════════════════════════════════════════
def step2_add_2026(folder: Path, log):
    log("\n── STEP 2: Add ' 2026' Suffix ───────────────────────────")
    renamed = skipped = 0
    for f in sorted(folder.iterdir()):
        if not f.is_file():
            continue
        if f.suffix.lower() not in (".xlsx", ".xls"):
            continue
        stem = f.stem
        if " 2026" in stem:
            log(f"  Skipped (already has ' 2026'): {f.name}")
            skipped += 1
            continue
        new_name = f"{stem} 2026{f.suffix}"
        dest = f.parent / new_name
        try:
            f.rename(dest)
            log(f"  Renamed: {f.name}  →  {new_name}")
            renamed += 1
        except Exception as e:
            log(f"  ERROR renaming {f.name}: {e}")
    log(f"  Done — {renamed} renamed, {skipped} skipped.")
    return renamed


# ═══════════════════════════════════════════════════════════════════════════
# STEP 3 — DELETE 2025 ROWS
# ═══════════════════════════════════════════════════════════════════════════
def _delete_2025_xlsx(path: Path, log) -> str:
    """Delete rows where col A contains '2025' from an .xlsx file."""
    if _load_xlsx is None:
        return "SKIP (openpyxl not installed)"
    wb = _load_xlsx(str(path))
    ws = wb.active
    rows_to_delete = [
        r for r in range(1, ws.max_row + 1)
        if "2025" in str(ws.cell(r, 1).value or "")
    ]
    if not rows_to_delete:
        return "SKIP (no 2025 rows)"
    # Delete bottom-up so row indices stay valid
    for r in reversed(rows_to_delete):
        ws.delete_rows(r)
    wb.save(str(path))
    return f"OK — deleted {len(rows_to_delete)} row(s)"


def _delete_2025_xls(path: Path, log) -> str:
    """
    Delete rows where col A contains '2025' from a legacy .xls file.
    Saves the result as .xlsx (modern format) so Excel / Power Query can open it.
    The original .xls file is removed after successful conversion.
    """
    if xlrd is None:
        return "SKIP (xlrd not installed)"
    if _load_xlsx is None:
        return "SKIP (openpyxl not installed)"

    # Read source with xlrd (only library that reads legacy .xls properly)
    rb = xlrd.open_workbook(str(path), formatting_info=False)
    rs = rb.sheet_by_index(0)

    keep_rows = [
        r for r in range(rs.nrows)
        if "2025" not in str(rs.cell_value(r, 0))
    ]
    deleted = rs.nrows - len(keep_rows)
    if deleted == 0:
        return "SKIP (no 2025 rows)"

    # Write kept rows into a new .xlsx workbook via openpyxl
    import openpyxl
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = rs.name

    for new_r_idx, old_r in enumerate(keep_rows, start=1):
        for c in range(rs.ncols):
            cell_val = rs.cell_value(old_r, c)
            # Convert xlrd floats that look like integers back to int
            if isinstance(cell_val, float) and cell_val == int(cell_val):
                cell_val = int(cell_val)
            new_ws.cell(row=new_r_idx, column=c + 1, value=cell_val)

    # Save as .xlsx next to the original .xls file
    xlsx_path = path.with_suffix(".xlsx")
    new_wb.save(str(xlsx_path))

    # Remove the old .xls so only the clean .xlsx remains
    try:
        path.unlink()
    except Exception:
        pass

    return f"OK — deleted {deleted} row(s), saved as {xlsx_path.name}"


def step3_delete_2025(folder: Path, log):
    log("\n── STEP 3: Delete 2025 Rows ─────────────────────────────")
    modified = skipped = errors = 0
    # Process both .xls (convert to .xlsx) and .xlsx (edit in-place)
    xls_files = [f for f in sorted(folder.iterdir())
                 if f.is_file() and f.suffix.lower() in (".xls", ".xlsx")]
    if not xls_files:
        log("  No Excel files found.")
        return 0
    for f in xls_files:
        log(f"  Processing: {f.name}")
        max_retries = 3
        for attempt in range(1, max_retries + 1):
            try:
                ext = f.suffix.lower()
                if ext == ".xlsx":
                    result = _delete_2025_xlsx(f, log)
                else:
                    result = _delete_2025_xls(f, log)
                log(f"    {result}")
                if result.startswith("OK"):
                    modified += 1
                else:
                    skipped += 1
                break
            except Exception as e:
                log(f"    Attempt {attempt}/{max_retries} failed: {e}")
                if attempt < max_retries:
                    time.sleep(2)
                else:
                    log(f"    FINAL FAILURE — skipping file.")
                    errors += 1
    log(f"  Done — {modified} modified, {skipped} skipped, {errors} errors.")
    return modified


# ═══════════════════════════════════════════════════════════════════════════
# GUI
# ═══════════════════════════════════════════════════════════════════════════
NAVY  = "#090d26"
RED   = "#e8212a"
WHITE = "#ffffff"
LIGHT = "#f4f4f4"

# ── GFH brand assets (icon + top-right logo) ───────────────────────────────
ICON_ICO_NAME = "gfh_icon.ico"
ICON_PNG_NAME = "gfh_icon.png"
WORDMARK_PNG_NAME = "gfh_wordmark.png"
COPYRIGHT_TEXT = "Created by Abad Umair Channa  |  Copyright © 2026  |  All rights reserved."

def get_script_dir():
    return (os.path.dirname(sys.executable)
            if getattr(sys, "frozen", False)
            else os.path.dirname(os.path.abspath(__file__)))


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("GFH Telecom — Rebate Folder Tools")
        self.geometry("780x600")
        self.resizable(True, True)
        self.configure(bg=NAVY)
        self._wordmark_img = None
        self._set_window_icon()
        self._build_ui()

    # ---- GFH branding: window icon (titlebar + taskbar) --------------------
    def _set_window_icon(self):
        icon_dir = get_script_dir()
        ico_path = os.path.join(icon_dir, ICON_ICO_NAME)
        png_path = os.path.join(icon_dir, ICON_PNG_NAME)
        try:
            if os.path.exists(ico_path):
                self.iconbitmap(ico_path)
        except Exception:
            pass
        try:
            if os.path.exists(png_path):
                self._icon_img = tk.PhotoImage(file=png_path)
                self.iconphoto(True, self._icon_img)
        except Exception:
            pass

    def _build_ui(self):
        # ── Header ──────────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=NAVY, height=54)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="GFH Telecom LLC — Rebate Folder Tools",
                 bg=NAVY, fg=WHITE, font=("Segoe UI", 13, "bold")).pack(
                 side="left", padx=18, pady=14)

        # ── Logo pinned to the TOP-RIGHT of the header ───────────────────────
        wordmark_path = os.path.join(get_script_dir(), WORDMARK_PNG_NAME)
        logo_holder = tk.Frame(hdr, bg=NAVY)
        logo_holder.pack(side="right", padx=(0, 18), pady=8)

        if os.path.exists(wordmark_path):
            try:
                self._wordmark_img = tk.PhotoImage(file=wordmark_path)
                tk.Label(logo_holder, image=self._wordmark_img, bg=NAVY).pack()
            except Exception:
                self._wordmark_img = None

        if self._wordmark_img is None:
            badge = tk.Label(
                logo_holder, text="G", bg=NAVY, fg=RED,
                font=("Segoe UI", 16, "bold"), width=2, highlightthickness=2,
                highlightbackground=RED, highlightcolor=RED,
            )
            badge.pack(side="left", padx=(0, 6))
            word_frame = tk.Frame(logo_holder, bg=NAVY)
            word_frame.pack(side="left")
            tk.Label(word_frame, text="GFH", bg=NAVY, fg=RED,
                     font=("Segoe UI", 11, "bold")).pack(anchor="w")
            tk.Label(word_frame, text="TELECOM", bg=NAVY, fg=WHITE,
                     font=("Segoe UI", 7, "bold")).pack(anchor="w")

        # ── Folder row ──────────────────────────────────────────────────────
        sep = tk.Frame(self, bg="#1e2340", height=1)
        sep.pack(fill="x")
        row = tk.Frame(self, bg=NAVY, pady=10)
        row.pack(fill="x", padx=18)
        tk.Label(row, text="Folder:", bg=NAVY, fg=WHITE,
                 font=("Segoe UI", 10)).pack(side="left")
        self.folder_var = tk.StringVar(value=(
            r"C:\Users\AbadUmairChanna\OneDrive - Verge Mobile\Documents\Rebate"
        ))
        tk.Entry(row, textvariable=self.folder_var, width=62,
                 font=("Segoe UI", 9)).pack(side="left", padx=8)
        tk.Button(row, text="Browse", bg="#1e2340", fg=WHITE,
                  font=("Segoe UI", 9), relief="flat", cursor="hand2",
                  command=self._browse).pack(side="left")

        # ── Step checkboxes ─────────────────────────────────────────────────
        opts = tk.Frame(self, bg=NAVY, pady=4)
        opts.pack(fill="x", padx=18)
        self.do_step1 = tk.BooleanVar(value=True)
        self.do_step2 = tk.BooleanVar(value=True)
        self.do_step3 = tk.BooleanVar(value=True)
        for var, label in [
            (self.do_step1, "1. Store Rename"),
            (self.do_step2, "2. Add ' 2026' Suffix"),
            (self.do_step3, "3. Delete 2025 Rows"),
        ]:
            tk.Checkbutton(opts, text=label, variable=var,
                           bg=NAVY, fg=WHITE, selectcolor=NAVY,
                           activebackground=NAVY, activeforeground=WHITE,
                           font=("Segoe UI", 10)).pack(side="left", padx=16)

        # ── Run button ──────────────────────────────────────────────────────
        btn_row = tk.Frame(self, bg=NAVY, pady=6)
        btn_row.pack(fill="x", padx=18)
        self.run_btn = tk.Button(btn_row, text="▶  Run", bg=RED, fg=WHITE,
                                  font=("Segoe UI", 11, "bold"), relief="flat",
                                  padx=22, pady=6, cursor="hand2",
                                  command=self._run)
        self.run_btn.pack(side="left")
        self.status_var = tk.StringVar(value="Ready.")
        tk.Label(btn_row, textvariable=self.status_var, bg=NAVY, fg="#aaaacc",
                 font=("Segoe UI", 9)).pack(side="left", padx=16)

        # ── Log area ────────────────────────────────────────────────────────
        log_frame = tk.Frame(self, bg=NAVY)
        log_frame.pack(fill="both", expand=True, padx=18, pady=(0, 6))
        tk.Label(log_frame, text="Log", bg=NAVY, fg=WHITE,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w")
        self.log_box = scrolledtext.ScrolledText(
            log_frame, wrap="word", font=("Consolas", 9),
            bg="#0d1230", fg="#d0d8ff", insertbackground=WHITE,
            relief="flat", height=18,
        )
        self.log_box.pack(fill="both", expand=True)

        # ── Copyright bar ───────────────────────────────────────────────────
        bar = tk.Frame(self, bg=NAVY, height=24)
        bar.pack(fill="x", side="bottom")
        bar.pack_propagate(False)
        tk.Label(bar, text=COPYRIGHT_TEXT, bg=NAVY, fg="#9d9db8",
                  font=("Segoe UI", 8)).pack(pady=3)

    def _browse(self):
        d = filedialog.askdirectory(title="Select Rebate folder")
        if d:
            self.folder_var.set(d)

    def _log(self, msg):
        self.log_box.insert("end", msg + "\n")
        self.log_box.see("end")
        self.update_idletasks()

    def _run(self):
        folder = Path(self.folder_var.get().strip())
        if not folder.exists():
            messagebox.showerror("Folder not found", str(folder))
            return
        self.run_btn.config(state="disabled")
        self.log_box.delete("1.0", "end")
        threading.Thread(target=self._worker, args=(folder,), daemon=True).start()

    def _worker(self, folder: Path):
        self.status_var.set("Running…")
        total = 0
        try:
            if self.do_step1.get():
                total += step1_store_rename(folder, self._log)
            if self.do_step2.get():
                total += step2_add_2026(folder, self._log)
            if self.do_step3.get():
                total += step3_delete_2025(folder, self._log)
            self._log(f"\n{'═'*52}")
            self._log(f"  All selected steps complete.")
            self.status_var.set("Done.")
        except Exception as e:
            self._log(f"\nCRITICAL ERROR: {e}")
            self.status_var.set("Error — see log.")
        finally:
            self.run_btn.config(state="normal")


if __name__ == "__main__":
    App().mainloop()
